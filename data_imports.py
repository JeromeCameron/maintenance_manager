"""
Import data from spreadsheets in data_files/ into the database.

Usage:
    uv run python data_imports.py
"""

import math
from datetime import date, datetime, time
from pathlib import Path

import openpyxl

from schema.database import engine
from schema.models import Asset, AssetCategory, AssetOwnership, AssetStatus, AssetSubStatus, AssetScores, WorkOrder, WorkOrderStatus, PurchaseOrder, PurchaseOrderType, Invoice, InvoiceStatus, InvoiceType, Downtime, Part, PartCategory, UnitMeasure, StockTransaction, StockLevel, TransactionType, EquipmentPart, Budget
from sqlmodel import Session, text


DATA_DIR = Path(__file__).parent / "data_files"


def _to_enum(enum_cls, value):
    """Return enum member for value, or None if value is None/unrecognised."""
    if value is None:
        return None
    try:
        return enum_cls(str(value).strip().lower())
    except ValueError:
        return None


def _to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value))
    except ValueError:
        return None


def import_assets(session: Session) -> None:
    path = DATA_DIR / "tbl_assets_pure.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        asset_id = data.get("asset_id")
        if not asset_id:
            skipped += 1
            continue

        asset = Asset(
            asset_id=str(asset_id).strip(),
            manufacturer=str(data["manufacturer"]).strip() if data.get("manufacturer") else "",
            model_no=str(data["model_no"]).strip() if data.get("model_no") else None,
            yr=int(data["yr"]) if data.get("yr") is not None else None,
            serial_no=str(data["serial_no"]).strip() if data.get("serial_no") else None,
            category=_to_enum(AssetCategory, data.get("category")),
            owned=_to_enum(AssetOwnership, data.get("owned")) or AssetOwnership.owned,
            alias=str(data["alias"]).strip() if data.get("alias") else None,
            date_in_service=_to_date(data.get("date_in_service")),
            status=_to_enum(AssetStatus, data.get("status")) or AssetStatus.operational,
            sub_status=_to_enum(AssetSubStatus, data.get("sub_status")),
            notes=str(data["notes"]).strip() if data.get("notes") else None,
            location_id=int(data["location"]) if data.get("location") is not None else None,
        )

        existing = session.get(Asset, asset.asset_id)
        session.merge(asset)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()
    print(f"Assets — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def _to_float(value) -> float | None:
    if value is None:
        return None
    try:
        f = float(value)
        return None if math.isnan(f) else f
    except (TypeError, ValueError):
        return None


def _to_int(value) -> int | None:
    f = _to_float(value)
    return None if f is None else int(f)


def _to_bool(value) -> bool | None:
    f = _to_float(value)
    if f is None:
        return None
    return bool(f)


_PRIORITY_MAP = {"urgent": "High", "high": "High", "medium": "Medium", "low": "Low"}
_TYPE_MAP = {"preventative": "preventative", "corrective": "corrective", "predictive": "predictive"}


def import_work_orders(session: Session) -> None:
    path = DATA_DIR / "tbl_workOrders.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    # Build sets of valid FK values to avoid constraint errors
    valid_assets = {row[0] for row in session.exec(text("SELECT asset_id FROM asset")).all()}  # type: ignore
    valid_suppliers = {row[0] for row in session.exec(text("SELECT supplier_id FROM supplier")).all()}  # type: ignore

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        wo_id = _to_int(data.get("work_order_id"))
        if wo_id is None:
            skipped += 1
            continue

        priority_raw = str(data.get("priority") or "").strip().lower()
        priority = _PRIORITY_MAP.get(priority_raw, "Low")

        typ_raw = str(data.get("typ") or "").strip().lower()
        typ = _TYPE_MAP.get(typ_raw, typ_raw)

        status_raw = str(data.get("status") or "").strip().lower()
        try:
            status = WorkOrderStatus(status_raw)
        except ValueError:
            status = WorkOrderStatus.requested

        asset_id = str(data["asset_id"]).strip() if data.get("asset_id") else None
        if asset_id and asset_id not in valid_assets:
            asset_id = None

        supplier_id = _to_int(data.get("supplier_id"))
        if supplier_id and supplier_id not in valid_suppliers:
            supplier_id = None

        wo = WorkOrder(
            work_order_id=wo_id,
            issue_date=_to_date(data.get("issue_date")),
            priority=priority,
            typ=typ,
            asset_id=asset_id,
            asset_pm_id=_to_int(data.get("asset_pm_id")),
            description=str(data["description"]).strip() if data.get("description") else None,
            supplier_id=supplier_id,
            expected_date=_to_date(data.get("expected_date")),
            date_completed=_to_date(data.get("date_completed")),
            estimated_cost=_to_float(data.get("estimated_cost")),
            estimated_hours=_to_float(data.get("estimated_hours")),
            actual_cost=_to_float(data.get("actual_cost")),
            actual_hours=_to_float(data.get("actual_hours")),
            notes=str(data["notes"]).strip() if data.get("notes") else None,
            status=status,
            planned=_to_bool(data.get("planned")),
        )

        existing = session.get(WorkOrder, wo_id)
        session.merge(wo)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()

    # Advance the sequence past the highest imported ID
    max_id = session.exec(text("SELECT MAX(work_order_id) FROM workorder")).first()[0]  # type: ignore
    if max_id:
        session.exec(text(f"ALTER SEQUENCE workorder_work_order_id_seq RESTART WITH {max_id + 1}"))  # type: ignore
        session.commit()

    print(f"Work Orders — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def import_purchase_orders(session: Session) -> None:
    path = DATA_DIR / "tbl_po.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    valid_assets = {row[0] for row in session.exec(text("SELECT asset_id FROM asset")).all()}  # type: ignore
    valid_suppliers = {row[0] for row in session.exec(text("SELECT supplier_id FROM supplier")).all()}  # type: ignore
    valid_locations = {row[0] for row in session.exec(text("SELECT location_id FROM location")).all()}  # type: ignore

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        po_no = str(data.get("po_no") or "").strip()
        if not po_no:
            skipped += 1
            continue

        po_type_raw = str(data.get("po_type") or "").strip().lower()
        try:
            po_type = PurchaseOrderType(po_type_raw)
        except ValueError:
            po_type = PurchaseOrderType.purchase

        asset_id = str(data["asset_id"]).strip() if data.get("asset_id") else None
        if asset_id and asset_id not in valid_assets:
            asset_id = None

        supplier_id = _to_int(data.get("supplier_id"))
        if supplier_id and supplier_id not in valid_suppliers:
            supplier_id = None

        location_id = _to_int(data.get("location_id"))
        if location_id and location_id not in valid_locations:
            location_id = None

        po = PurchaseOrder(
            po_no=po_no,
            po_date=_to_date(data.get("po_date")),
            supplier_id=supplier_id,
            asset_id=asset_id,
            location_id=location_id,
            description=str(data["description"]).strip() if data.get("description") else None,
            subtotal=_to_float(data.get("subtotal")) or 0.0,
            po_type=po_type,
            cost_centre_id=None,
        )

        existing = session.get(PurchaseOrder, po_no)
        session.merge(po)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()
    print(f"Purchase Orders — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def import_invoices(session: Session) -> None:
    path = DATA_DIR / "tbl_invoices.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    valid_suppliers = {row[0] for row in session.exec(text("SELECT supplier_id FROM supplier")).all()}  # type: ignore
    valid_assets = {row[0] for row in session.exec(text("SELECT asset_id FROM asset")).all()}  # type: ignore
    valid_locations = {row[0] for row in session.exec(text("SELECT location_id FROM location")).all()}  # type: ignore

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        inv_id = _to_int(data.get("ID"))
        if inv_id is None:
            skipped += 1
            continue

        invoice_no = str(data.get("invoice_no") or "").strip()
        if not invoice_no:
            skipped += 1
            continue

        try:
            status = InvoiceStatus(str(data.get("status") or "").strip().lower())
        except ValueError:
            status = InvoiceStatus.submitted

        try:
            invoice_type = InvoiceType(str(data.get("invoice_type") or "").strip().lower())
        except ValueError:
            invoice_type = InvoiceType.services

        supplier_id = _to_int(data.get("supplier_id"))
        if supplier_id and supplier_id not in valid_suppliers:
            supplier_id = None

        asset_id = str(data["asset_id"]).strip() if data.get("asset_id") else None
        if asset_id and asset_id not in valid_assets:
            asset_id = None

        location_id = _to_int(data.get("location_id"))
        if location_id and location_id not in valid_locations:
            location_id = None

        tax_cert_val = data.get("tax_cert")
        if tax_cert_val is None:
            tax_cert = None
        elif isinstance(tax_cert_val, bool):
            tax_cert = tax_cert_val
        else:
            tax_cert = _to_bool(tax_cert_val)

        inv = Invoice(
            id=inv_id,
            invoice_no=invoice_no,
            invoice_date=_to_date(data.get("invoice_date")),
            job_date=_to_date(data.get("job_date")),
            rec_date=_to_date(data.get("rec_date")),
            supplier_id=supplier_id,
            work_order_id=None,
            asset_id=asset_id,
            location_id=location_id,
            description=str(data["description"]).strip() if data.get("description") else None,
            po_no=None,
            subtotal=_to_float(data.get("subtotal")) or 0.0,
            status=status,
            tax_cert=tax_cert,
            invoice_type=invoice_type,
        )

        existing = session.get(Invoice, inv_id)
        session.merge(inv)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()

    max_id = session.exec(text("SELECT MAX(id) FROM invoice")).first()[0]  # type: ignore
    if max_id:
        session.exec(text(f"ALTER SEQUENCE invoice_id_seq RESTART WITH {max_id + 1}"))  # type: ignore
        session.commit()

    print(f"Invoices — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def import_downtimes(session: Session) -> None:
    path = DATA_DIR / "tbl_downtime.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    valid_assets = {row[0] for row in session.exec(text("SELECT asset_id FROM asset")).all()}  # type: ignore
    valid_causes = {row[0] for row in session.exec(text("SELECT cause_id FROM downtimecause")).all()}  # type: ignore

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        downtime_id = _to_int(data.get("downtime_id"))
        if downtime_id is None:
            skipped += 1
            continue

        log_date_val = data.get("log_date")
        log_time_val = data.get("log_time")
        if isinstance(log_date_val, datetime):
            log_dt = log_date_val.replace(
                hour=log_time_val.hour if isinstance(log_time_val, time) else 0,
                minute=log_time_val.minute if isinstance(log_time_val, time) else 0,
                second=0,
                microsecond=0,
            )
        else:
            log_dt = None

        start_date_val = data.get("start_date")
        start_date = start_date_val.date() if isinstance(start_date_val, datetime) else _to_date(start_date_val)

        end_date_val = data.get("end_date")
        end_date = end_date_val.date() if isinstance(end_date_val, datetime) else _to_date(end_date_val)

        asset_id = str(data["asset_id"]).strip() if data.get("asset_id") else None
        if asset_id and asset_id not in valid_assets:
            asset_id = None

        cause_id = _to_int(data.get("cause_id"))
        if cause_id and cause_id not in valid_causes:
            cause_id = None

        dt = Downtime(
            downtime_id=downtime_id,
            log_date=log_dt,
            asset_id=asset_id,
            shift_asset=bool(data.get("shift_asset")),
            cause_id=cause_id,
            start_date=start_date,
            start_time=data.get("start_time") if isinstance(data.get("start_time"), time) else None,
            end_date=end_date,
            end_time=data.get("end_time") if isinstance(data.get("end_time"), time) else None,
            planned=_to_bool(data.get("planned")),
            details=str(data["details"]).strip() if data.get("details") else None,
            component_affected=str(data["component_affected"]).strip() if data.get("component_affected") else None,
            root_cause=str(data["root_cause"]).strip() if data.get("root_cause") else None,
            corrective_action=str(data["corrective_action"]).strip() if data.get("corrective_action") else None,
            repeat_failure=_to_bool(data.get("repeat_failure")),
            temporary_fix=_to_bool(data.get("temporary_fix")),
            work_order=str(data["work_order"]).strip() if data.get("work_order") else "",
            downtime_hours=_to_float(data.get("downtime_hours")),
        )

        existing = session.get(Downtime, downtime_id)
        session.merge(dt)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()

    max_id = session.exec(text("SELECT MAX(downtime_id) FROM downtime")).first()[0]  # type: ignore
    if max_id:
        session.exec(text(f"ALTER SEQUENCE downtime_downtime_id_seq RESTART WITH {max_id + 1}"))  # type: ignore
        session.commit()

    print(f"Downtimes — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def import_parts(session: Session) -> None:
    path = DATA_DIR / "tbl_parts.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    # Cache existing categories and create missing ones
    category_map: dict[str, int] = {}
    for row in session.exec(text("SELECT id, name FROM partcategory")).all():  # type: ignore
        category_map[row[1]] = row[0]

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        part_no = str(data.get("part_no") or "").strip()
        if not part_no:
            skipped += 1
            continue

        # Resolve or create category
        category_name = str(data.get("category") or "").strip()
        category_id: int | None = None
        if category_name:
            if category_name not in category_map:
                cat = PartCategory(name=category_name)
                session.add(cat)
                session.flush()
                category_map[category_name] = cat.id  # type: ignore
            category_id = category_map[category_name]

        unit_raw = str(data.get("unit_of_measure") or "").strip().lower()
        unit = _to_enum(UnitMeasure, unit_raw)

        part = Part(
            part_no=part_no,
            part_name=str(data["part_name"]).strip() if data.get("part_name") else "",
            manufacturer=str(data["manufacturer"]).strip() if data.get("manufacturer") else None,
            description=str(data["description"]).strip() if data.get("description") else None,
            category_id=category_id,
            unit_of_measure=unit,
            min_level=_to_int(data.get("min_level")) or 0,
            max_level=_to_int(data.get("max_level")) or 0,
            reorder_level=_to_int(data.get("reorder_level")) or 0,
            reorder_qty=_to_int(data.get("reorder_qty")) or 0,
            last_cost=_to_float(data.get("last_cost")),
            is_critical=bool(data.get("is_critical")),
            is_active=bool(data.get("active")) if data.get("active") is not None else True,
        )

        existing = session.get(Part, part_no)
        session.merge(part)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()
    print(f"Parts — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def import_stock_transactions(session: Session) -> None:
    path = DATA_DIR / "tbl_StockTransactions.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    valid_parts = {row[0] for row in session.exec(text("SELECT part_no FROM part")).all()}  # type: ignore
    valid_locations = {row[0] for row in session.exec(text("SELECT location_id FROM location")).all()}  # type: ignore
    valid_work_orders = {row[0] for row in session.exec(text("SELECT work_order_id FROM workorder")).all()}  # type: ignore
    valid_po = {row[0] for row in session.exec(text("SELECT po_no FROM purchaseorder")).all()}  # type: ignore

    # Build name → user_id map
    user_map: dict[str, int] = {}
    for row in session.exec(text('SELECT id, firstname, lastname FROM "user"')).all():  # type: ignore
        user_map[f"{row[1]} {row[2]}"] = row[0]

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        txn_id = _to_int(data.get("id"))
        if txn_id is None:
            skipped += 1
            continue

        part_no = str(data.get("part_id") or "").strip()
        if part_no not in valid_parts:
            part_no = None  # type: ignore

        location_id = _to_int(data.get("location_id"))
        if location_id and location_id not in valid_locations:
            location_id = None

        work_order_id = _to_int(data.get("work_order_id"))
        if work_order_id and work_order_id not in valid_work_orders:
            work_order_id = None

        po_no = str(data.get("po_no") or "").strip() or None
        if po_no and po_no not in valid_po:
            po_no = None

        entered_by_name = str(data.get("entered_by") or "").strip()
        entered_by = user_map.get(entered_by_name)

        txn_date = data.get("transaction_date")
        if not isinstance(txn_date, datetime):
            txn_date = None

        txn = StockTransaction(
            id=txn_id,
            part_no=part_no,
            location_id=location_id,
            transaction_type=_to_enum(TransactionType, str(data.get("transaction_type") or "").lower()),
            quantity=_to_int(data.get("quanity")) or 0,
            transaction_date=txn_date or datetime.now(),
            work_order_id=work_order_id,
            po_no=po_no,
            entered_by=entered_by,
            notes=str(data["notes"]).strip() if data.get("notes") else None,
        )

        existing = session.get(StockTransaction, txn_id)
        session.merge(txn)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()

    max_id = session.exec(text("SELECT MAX(id) FROM stocktransaction")).first()[0]  # type: ignore
    if max_id:
        session.exec(text(f"ALTER SEQUENCE stocktransaction_id_seq RESTART WITH {max_id + 1}"))  # type: ignore
        session.commit()

    print(f"Stock Transactions — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def import_equipment_parts(session: Session) -> None:
    path = DATA_DIR / "tbl_EquipmentParts.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    valid_models = {row[0] for row in session.exec(text("SELECT model_no FROM assetmodel")).all()}  # type: ignore
    valid_parts = {row[0] for row in session.exec(text("SELECT part_no FROM part")).all()}  # type: ignore

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        ep_id = _to_int(data.get("ID"))
        if ep_id is None:
            skipped += 1
            continue

        model_no = str(data.get("model_no") or "").strip() or None
        if model_no and model_no not in valid_models:
            skipped += 1
            continue

        part_no = str(data.get("part_id") or "").strip() or None
        if part_no and part_no not in valid_parts:
            skipped += 1
            continue

        ep = EquipmentPart(
            id=ep_id,
            model_no=model_no,
            part_no=part_no,
            is_critical=bool(data.get("is_critical")),
        )

        existing = session.get(EquipmentPart, ep_id)
        session.merge(ep)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()

    max_id = session.exec(text("SELECT MAX(id) FROM equipmentpart")).first()[0]  # type: ignore
    if max_id:
        session.exec(text(f"ALTER SEQUENCE equipmentpart_id_seq RESTART WITH {max_id + 1}"))  # type: ignore
        session.commit()

    print(f"Equipment Parts — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def import_budget(session: Session) -> None:
    path = DATA_DIR / "tbl_budget.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    valid_gl_codes = {row[0] for row in session.exec(text("SELECT gl_code FROM costcentre")).all()}  # type: ignore

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        budget_id = _to_int(data.get("ID"))
        if budget_id is None:
            skipped += 1
            continue

        gl_code = str(data.get("gl_code") or "").strip() or None
        if gl_code and gl_code not in valid_gl_codes:
            gl_code = None

        mn_val = data.get("mn")
        if isinstance(mn_val, datetime):
            month = mn_val.date()
        else:
            month = _to_date(mn_val)

        if month is None:
            skipped += 1
            continue

        budget = Budget(
            id=budget_id,
            gl_code=gl_code,
            financial_year=str(data.get("financial_year") or "").strip(),
            month=month,
            amount=_to_float(data.get("amount")) or 0.0,
            notes=str(data["notes"]).strip() if data.get("notes") else None,
        )

        existing = session.get(Budget, budget_id)
        session.merge(budget)

        if existing:
            updated += 1
        else:
            inserted += 1

    session.commit()

    max_id = session.exec(text("SELECT MAX(id) FROM budget")).first()[0]  # type: ignore
    if max_id:
        session.exec(text(f"ALTER SEQUENCE budget_id_seq RESTART WITH {max_id + 1}"))  # type: ignore
        session.commit()

    print(f"Budget — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


def rebuild_stock_levels(session: Session) -> None:
    """Recalculate StockLevel for every (part_no, location_id) from all transactions."""
    rows = session.exec(
        text("""
            SELECT part_no, location_id,
                   SUM(CASE WHEN transaction_type = 'receive' THEN quantity
                            WHEN transaction_type = 'issue'   THEN -quantity
                            ELSE quantity END) AS net_qty
            FROM stocktransaction
            WHERE part_no IS NOT NULL
            GROUP BY part_no, location_id
        """)
    ).all()  # type: ignore

    upserted = 0
    for part_no, location_id, net_qty in rows:
        existing = session.exec(
            text("SELECT id FROM stocklevel WHERE part_no = :p AND (location_id = :l OR (location_id IS NULL AND :l IS NULL))"),
            params={"p": part_no, "l": location_id},  # type: ignore
        ).first()

        if existing:
            level = session.get(StockLevel, existing[0])
            level.quantity = int(net_qty or 0)  # type: ignore
            level.last_updated = datetime.now()  # type: ignore
        else:
            session.add(StockLevel(
                part_no=part_no,
                location_id=location_id,
                quantity=int(net_qty or 0),
                last_updated=datetime.now(),
            ))
        upserted += 1

    session.commit()
    print(f"Stock Levels — upserted: {upserted}")


def import_criticality_scores(session: Session) -> None:
    path = DATA_DIR / "qry_criticality_scores.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    valid_assets = {row[0] for row in session.exec(text("SELECT asset_id FROM asset")).all()}  # type: ignore

    inserted = updated = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        asset_id = str(data.get("asset_id") or "").strip()
        if not asset_id or asset_id not in valid_assets:
            skipped += 1
            continue

        existing = session.exec(
            text("SELECT score_id FROM assetscores WHERE asset_id = :aid"),
            params={"aid": asset_id},  # type: ignore
        ).first()

        if existing:
            score_id = existing[0]
            rec = session.get(AssetScores, score_id)
            rec.operational_score = _to_int(data.get("operational_score"))  # type: ignore
            rec.safety_score = _to_int(data.get("safety_score"))  # type: ignore
            rec.backup_score = _to_int(data.get("backup_score"))  # type: ignore
            rec.repair_score = _to_int(data.get("repair_score"))  # type: ignore
            rec.usage_score = _to_int(data.get("usage_score"))  # type: ignore
            updated += 1
        else:
            session.add(AssetScores(
                asset_id=asset_id,
                operational_score=_to_int(data.get("operational_score")),
                safety_score=_to_int(data.get("safety_score")),
                backup_score=_to_int(data.get("backup_score")),
                repair_score=_to_int(data.get("repair_score")),
                usage_score=_to_int(data.get("usage_score")),
            ))
            inserted += 1

    session.commit()
    print(f"Criticality Scores — inserted: {inserted}, updated: {updated}, skipped: {skipped}")


if __name__ == "__main__":
    with Session(engine) as session:
        import_assets(session)
        import_work_orders(session)
        import_purchase_orders(session)
        import_invoices(session)
        import_downtimes(session)
        import_criticality_scores(session)
        import_parts(session)
        import_stock_transactions(session)
        rebuild_stock_levels(session)
        import_equipment_parts(session)
        import_budget(session)
